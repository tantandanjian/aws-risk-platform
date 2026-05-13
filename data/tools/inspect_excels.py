from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "aws-risk-platform" / "data" / "_excel_inspection.json"


def clean(value):
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return " ".join(text.split())


def sheet_snapshot(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = [clean(v) for v in row]
        if any(values):
            rows.append(values)

    header_idx = 0
    best_score = -1
    for idx, values in enumerate(rows[:20]):
        score = sum(1 for v in values if v)
        joined = " ".join(values).lower()
        if any(token in joined for token in ["api", "风险", "配置", "service", "接口", "字段", "说明"]):
            score += 5
        if score > best_score:
            best_score = score
            header_idx = idx

    headers = rows[header_idx] if rows else []
    sample_rows = rows[header_idx + 1 : header_idx + 11] if rows else []
    return {
        "sheet": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "non_empty_rows": len(rows),
        "header_row_1_based": header_idx + 1 if rows else None,
        "headers": headers,
        "sample_rows": sample_rows,
    }


def main():
    files = []
    for path in sorted(ROOT.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        files.append(
            {
                "file": path.name,
                "sheets": [sheet_snapshot(ws) for ws in wb.worksheets],
            }
        )
        wb.close()

    OUT.write_text(json.dumps(files, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} ({len(files)} workbooks)")


if __name__ == "__main__":
    main()
