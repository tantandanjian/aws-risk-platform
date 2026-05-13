from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
SITE = Path(__file__).resolve().parents[1]
DATA_DIR = SITE / "data"


SERVICE_META = {
    "cloudfront": {
        "service": "Amazon CloudFront",
        "aws_category": "Networking & Content Delivery",
        "service_group": "TLS / 安全策略",
        "crypto_category": "TLS/HTTPS; Certificate",
        "configurable_status": "可配置",
        "service_intro": "CloudFront 是 AWS 的内容分发网络服务，支持 HTTPS、TLS 安全策略和证书配置。其密码相关配置主要集中在 Viewer TLS 策略、证书绑定和 HTTPS 访问控制等方面。",
    },
    "alb": {
        "service": "AWS Application Load Balancer",
        "aws_category": "Networking & Content Delivery",
        "service_group": "TLS / 安全策略",
        "crypto_category": "TLS/HTTPS; Certificate",
        "configurable_status": "可配置",
        "service_intro": "Application Load Balancer 可通过 HTTPS listener 暴露应用入口，密码相关风险主要集中在 listener SSL policy、证书绑定和旧 TLS 策略。",
    },
    "nlb": {
        "service": "AWS Network Load Balancer",
        "aws_category": "Networking & Content Delivery",
        "service_group": "TLS / 安全策略",
        "crypto_category": "TLS/HTTPS; Certificate",
        "configurable_status": "可配置",
        "service_intro": "Network Load Balancer 支持 TLS listener，相关密码配置集中在 TLS 安全策略、证书绑定和旧协议兼容。",
    },
    "vpc-lattice": {
        "service": "Amazon VPC Lattice",
        "aws_category": "Networking & Content Delivery",
        "service_group": "TLS / 安全策略",
        "crypto_category": "TLS/HTTPS; Certificate",
        "configurable_status": "部分可配置",
        "service_intro": "VPC Lattice 的密码相关配置主要通过 listener、custom domain 和证书绑定体现，底层 TLS 细节通常由 AWS 托管。",
    },
    "route-53": {
        "service": "Amazon Route 53",
        "aws_category": "Networking & Content Delivery",
        "service_group": "DNSSEC / 完整性",
        "crypto_category": "DNSSEC",
        "configurable_status": "可配置",
        "service_intro": "Route 53 的密码相关能力主要体现在 DNSSEC signing，用于增强 DNS 响应完整性和来源认证。",
    },
    "kms": {
        "service": "AWS KMS",
        "aws_category": "Security, Identity, & Compliance",
        "service_group": "密钥管理",
        "crypto_category": "KMS; KeySpec; KeyUsage; Rotation",
        "configurable_status": "可配置",
        "service_intro": "AWS KMS 用于创建和管理加密、解密、签名等用途的密钥，重点关注 KeySpec、KeyUsage、Origin、轮换和访问控制。",
    },
    "cloudhsm": {
        "service": "AWS CloudHSM",
        "aws_category": "Security, Identity, & Compliance",
        "service_group": "HSM / 密码机制",
        "crypto_category": "HSM; PKCS#11; JCE; CNG",
        "configurable_status": "部分可配置",
        "service_intro": "CloudHSM 提供云中托管 HSM，密码风险主要来自机制选择、密钥管理方式和遗留算法使用。",
    },
    "acm": {
        "service": "AWS Certificate Manager",
        "aws_category": "Security, Identity, & Compliance",
        "service_group": "证书管理",
        "crypto_category": "Certificate; TLS",
        "configurable_status": "可配置",
        "service_intro": "ACM 用于申请、导入、存储、续期和部署 SSL/TLS X.509 证书，重点关注证书来源、密钥算法、验证方式和续期状态。",
    },
    "signer": {
        "service": "AWS Signer",
        "aws_category": "Security, Identity, & Compliance",
        "service_group": "数字签名",
        "crypto_category": "Digital Signature; Code Signing",
        "configurable_status": "部分可配置",
        "service_intro": "AWS Signer 提供托管代码签名能力，重点关注 signing profile、platformId、签名证书和权限治理。",
    },
    "encryption-sdk": {
        "service": "AWS Encryption SDK",
        "aws_category": "Developer Tools",
        "service_group": "客户端加密 SDK",
        "crypto_category": "Client-side Encryption; Keyring; Commitment Policy",
        "configurable_status": "可配置",
        "service_intro": "AWS Encryption SDK 面向应用侧加密，重点关注 keyring、wrapping keys、commitment policy、algorithm suite 和 encryption context。",
    },
    "database-encryption-sdk": {
        "service": "AWS Database Encryption SDK",
        "aws_category": "Developer Tools",
        "service_group": "客户端加密 SDK",
        "crypto_category": "Client-side Encryption; Searchable Encryption",
        "configurable_status": "可配置",
        "service_intro": "AWS Database Encryption SDK 面向数据库字段级客户端加密，重点关注 keyring、wrapping key、加密上下文和可搜索加密 beacon 设计。",
    },
    "s3-encryption-client": {
        "service": "Amazon S3 Encryption Client",
        "aws_category": "Storage",
        "service_group": "客户端加密 SDK",
        "crypto_category": "Client-side Encryption; Keyring",
        "configurable_status": "可配置",
        "service_intro": "Amazon S3 Encryption Client 用于 S3 客户端加密，重点关注 wrapping key、keyring、材料管理和原始密钥使用风险。",
    },
    "payment-cryptography": {
        "service": "AWS Payment Cryptography",
        "aws_category": "Financial Services",
        "service_group": "支付密码",
        "crypto_category": "Payment Cryptography; PIN; CVV; TR-31; TR-34",
        "configurable_status": "可配置",
        "service_intro": "AWS Payment Cryptography 面向支付场景的密钥和密码运算，重点关注密钥类别、算法、用途、导入导出和允许操作。",
    },
    "vpc": {
        "service": "Amazon VPC",
        "aws_category": "Networking & Content Delivery",
        "service_group": "不可直接配置",
        "crypto_category": "N/A",
        "configurable_status": "不可直接配置",
        "service_intro": "Amazon VPC 本身不暴露密码套件、TLS policy 或算法选择入口，安全能力主要体现为网络隔离、访问控制和上层服务加密配置承载。",
    },
    "privatelink": {
        "service": "AWS PrivateLink",
        "aws_category": "Networking & Content Delivery",
        "service_group": "不可直接配置",
        "crypto_category": "N/A",
        "configurable_status": "不可直接配置",
        "service_intro": "AWS PrivateLink 本身不提供直接密码算法选择，通常通过私有连接和上层服务 TLS/证书配置共同实现安全通信。",
    },
    "cloud-map": {
        "service": "AWS Cloud Map",
        "aws_category": "Networking & Content Delivery",
        "service_group": "不可直接配置",
        "crypto_category": "N/A",
        "configurable_status": "不可直接配置",
        "service_intro": "AWS Cloud Map 主要用于服务发现，不直接暴露密码配置入口；TLS、证书和加密通常在调用方、服务端或上层网关中配置。",
    },
}


RISK_TYPES = [
    ("old-tls", "TLS 旧版本风险", "使用 SSLv3、TLS 1.0、TLS 1.1 或旧安全策略，可能启用过时协议和弱密码套件。", "CloudFront; ALB; NLB", "SSLv3; TLSv1; TLSv1.1", "使用 TLS 1.2 / TLS 1.3 和较新的安全策略"),
    ("old-security-policy", "TLS 安全策略过旧", "使用较早的 ELB / CloudFront security policy，可能包含非前向安全或旧式密码套件。", "CloudFront; ALB; NLB", "ELBSecurityPolicy-2016-08; TLSv1_2016", "选择现代 TLS 1.2/1.3 policy"),
    ("certificate", "证书管理风险", "使用过期证书、默认证书、弱密钥证书或证书链配置错误。", "ACM; CloudFront; ALB; NLB", "过期证书; 默认证书; RSA_1024", "使用 ACM 托管证书、有效自定义证书和自动续期"),
    ("dnssec", "DNSSEC 未启用风险", "重要 public hosted zone 未启用 DNSSEC，DNS 响应缺少完整性保护。", "Route 53", "DNSSEC signing disabled", "启用 Route 53 DNSSEC signing"),
    ("kms", "KMS 密钥配置风险", "KeySpec、KeyUsage、Origin 等配置与业务用途不匹配。", "AWS KMS", "用途混用; key spec 不匹配", "按用途选择 KeySpec、KeyUsage 和 Origin"),
    ("lifecycle", "密钥生命周期风险", "密钥轮换关闭、证书未续期或外部密钥材料生命周期管理不足。", "KMS; ACM", "未启用轮换; 导入证书未续期", "启用轮换和托管续期"),
    ("hsm", "HSM 机制风险", "使用遗留或不推荐的密码机制。", "CloudHSM", "旧机制; 弱哈希", "使用现代密码机制并限制遗留兼容"),
    ("signing", "数字签名治理风险", "签名 profile、权限或签名流程管理不当。", "AWS Signer", "SHA1; 弱签名平台", "使用现代签名平台和受控 signing profile"),
    ("sdk-policy", "SDK 加密策略风险", "commitment policy、algorithm suite、keyring 配置不当。", "AWS Encryption SDK; S3 Encryption Client", "裸 discovery keyring; raw key 管理不当", "使用受控 KMS keyring 和推荐 commitment policy"),
    ("encryption-context", "加密上下文缺失风险", "encryption context 缺失或缺乏业务语义绑定。", "AWS Encryption SDK; Database Encryption SDK", "空上下文; 无业务语义", "绑定有业务语义的 encryption context"),
    ("searchable-encryption", "可搜索加密泄露风险", "beacon 设计过宽，可能扩大查询模式泄露面。", "Database Encryption SDK", "过宽 beacon", "最小化 beacon 设计并评估泄露面"),
    ("payment-key", "支付密码密钥管理风险", "支付密钥用途、算法、导入导出和访问控制配置不当。", "AWS Payment Cryptography", "TDES_2KEY; 用途混用; exportable 过宽", "按 TR-31/TR-34 场景限制密钥用途和操作"),
    ("not-configurable", "不可直接配置说明", "服务本身不暴露密码套件、TLS policy 或算法选择。", "VPC; PrivateLink; Cloud Map", "N/A", "在上层服务或应用层配置 TLS、证书、加密和访问控制"),
]


def text(value) -> str:
    if value is None:
        return ""
    value = str(value).replace("\xa0", " ").strip()
    value = re.sub(r"\s+", " ", value)
    return value


def slug(value: str) -> str:
    value = value.lower().replace("&", " and ")
    value = re.sub(r"aws|amazon", "", value)
    value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
    aliases = {
        "application-load-balancer": "alb",
        "alb": "alb",
        "network-load-balancer": "nlb",
        "nlb": "nlb",
        "cloudfront": "cloudfront",
        "route-53": "route-53",
        "kms": "kms",
        "key-management-service": "kms",
        "cloudhsm": "cloudhsm",
        "certificate-manager": "acm",
        "acm": "acm",
        "signer": "signer",
        "encryption-sdk": "encryption-sdk",
        "database-encryption-sdk": "database-encryption-sdk",
        "s3-encryption-client": "s3-encryption-client",
        "payment-cryptography": "payment-cryptography",
        "vpc": "vpc",
        "privatelink": "privatelink",
        "private-link": "privatelink",
        "cloud-map": "cloud-map",
    }
    return aliases.get(value, value or "service")


def service_id_from_filename(path: Path) -> str | None:
    name = path.name.lower()
    if "payment" in name:
        return "payment-cryptography"
    if "database_encryption_sdk" in name or "database encryption sdk" in name:
        return "database-encryption-sdk"
    if "s3_encryption_client" in name or "s3 encryption client" in name:
        return "s3-encryption-client"
    if "encryption sdk" in name and "database" not in name:
        return "encryption-sdk"
    if "signer" in name:
        return "signer"
    if "acm" in name:
        return "acm"
    return None


def infer_risk_type(service_id: str, row_text: str) -> tuple[str, str]:
    s = row_text.lower()
    if service_id in {"cloudfront", "alb", "nlb"} or "tls" in s or "ssl" in s:
        if "certificate" in s or "证书" in s:
            return "certificate", "证书管理风险"
        return "old-tls", "TLS 旧版本风险"
    if "dnssec" in s:
        return "dnssec", "DNSSEC 未启用风险"
    if service_id == "kms" or "keyspec" in s or "keyusage" in s or "rotation" in s:
        return "kms", "KMS 密钥配置风险"
    if service_id == "cloudhsm" or "hsm" in s or "pkcs" in s:
        return "hsm", "HSM 机制风险"
    if service_id == "acm" or "certificate" in s or "证书" in s:
        return "certificate", "证书管理风险"
    if service_id == "signer" or "sign" in s or "签名" in s:
        return "signing", "数字签名治理风险"
    if "context" in s or "上下文" in s:
        return "encryption-context", "加密上下文缺失风险"
    if "beacon" in s or "search" in s or "搜索" in s:
        return "searchable-encryption", "可搜索加密泄露风险"
    if service_id in {"encryption-sdk", "database-encryption-sdk", "s3-encryption-client"}:
        return "sdk-policy", "SDK 加密策略风险"
    if service_id == "payment-cryptography" or "tr-31" in s or "pin" in s or "cvv" in s:
        return "payment-key", "支付密码密钥管理风险"
    if service_id in {"vpc", "privatelink", "cloud-map"}:
        return "not-configurable", "不可直接配置说明"
    return "old-security-policy", "TLS 安全策略过旧"


def risk_level(risk_type_id: str, risky: str) -> str:
    risky_l = risky.lower()
    if risk_type_id in {"old-tls", "payment-key"} or any(x in risky_l for x in ["sslv3", "tlsv1", "rsa_1024", "sha1", "tdes_2key"]):
        return "高"
    if risk_type_id in {"certificate", "kms", "sdk-policy", "searchable-encryption"}:
        return "中"
    if risk_type_id == "not-configurable":
        return "说明性"
    return "中"


def infer_crypto_category(service_id: str, value: str = "") -> str:
    if value:
        return value
    return SERVICE_META.get(service_id, {}).get("crypto_category", "Cryptography")


def add_item(items: list[dict], raw: dict):
    config = text(raw.get("configuration_item"))
    service_id = raw["service_id"]
    if not config:
        return
    row_text = " ".join(text(v) for v in raw.values())
    risk_type_id, risk_type_name = infer_risk_type(service_id, row_text)
    item_id = f"{service_id}_{slug(config)[:48]}"
    item = {
        "item_id": item_id,
        "service_id": service_id,
        "configuration_item": config,
        "api_endpoint": text(raw.get("api_endpoint")) or "N/A",
        "recommended_values": text(raw.get("recommended_values")) or "N/A",
        "risky_values": text(raw.get("risky_values")) or "N/A",
        "risk_type": text(raw.get("risk_type")) or risk_type_name,
        "risk_level": text(raw.get("risk_level")) or risk_level(risk_type_id, text(raw.get("risky_values"))),
        "risk_reason": text(raw.get("risk_reason")) or f"该配置项与“{risk_type_name}”相关，风险值可能导致密码能力降级、治理边界变宽或与业务用途不匹配。",
        "security_value_reason": text(raw.get("security_value_reason")) or "推荐值更符合现代密码配置基线，有助于减少弱协议、弱算法、密钥误用或生命周期管理不足带来的风险。",
        "references": text(raw.get("references")) or "AWS 官方文档",
    }
    items.append(item)


def extract_unified_summary(path: Path) -> list[dict]:
    items: list[dict] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = [[text(v) for v in row] for row in ws.iter_rows(values_only=True)]
        header_idx = None
        for i, row in enumerate(rows):
            if "Public_Cloud" in row and "Service" in row:
                header_idx = i
                break
        if header_idx is None:
            continue
        headers = rows[header_idx]
        for row in rows[header_idx + 1 :]:
            if not any(row):
                continue
            rec = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
            service = text(rec.get("Service"))
            if not service or service.lower() == "service":
                continue
            sid = slug(service)
            config = text(rec.get("Configuration_Item")) or text(rec.get("API_Endpoint")) or "密码相关配置项"
            add_item(items, {
                "service_id": sid,
                "configuration_item": config,
                "api_endpoint": rec.get("API_Endpoint", ""),
                "recommended_values": rec.get("Recommended_Values", ""),
                "risky_values": rec.get("Risky_Values", ""),
                "risk_reason": rec.get("Risk_Reason", ""),
                "security_value_reason": rec.get("Security_Value_Reason", ""),
                "references": rec.get("References", ""),
            })
    return items


def find_header(row: list[str]) -> bool:
    joined = "|".join(row)
    return ("配置项" in joined and "风险值" in joined and ("安全值" in joined or "推荐值" in joined))


def extract_special_workbook(path: Path) -> list[dict]:
    service_id = service_id_from_filename(path)
    if not service_id:
        return []
    items: list[dict] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = [[text(v) for v in row] for row in ws.iter_rows(values_only=True)]
        header_idx = None
        for i, row in enumerate(rows[:8]):
            if find_header(row):
                header_idx = i
                break
        if header_idx is None:
            continue
        headers = rows[header_idx]
        normalized = []
        for h in headers:
            if "配置项" in h:
                normalized.append("configuration_item")
            elif "可选值" in h or "典型值" in h:
                normalized.append("options")
            elif "风险值" in h:
                normalized.append("risky_values")
            elif "安全值" in h or "推荐值" in h:
                normalized.append("recommended_values")
            elif "命令" in h or "API" in h or "对应位置" in h:
                normalized.append("api_endpoint")
            elif "参考" in h or "依据" in h:
                normalized.append("references")
            elif "原因" in h:
                normalized.append("risk_reason")
            else:
                normalized.append("")
        for row in rows[header_idx + 1 :]:
            if not any(row) or not row[0]:
                continue
            rec = {"service_id": service_id}
            for i, key in enumerate(normalized):
                if key:
                    rec[key] = row[i] if i < len(row) else ""
            if "options" in rec and rec.get("api_endpoint"):
                rec["api_endpoint"] = f"{rec['api_endpoint']}"
            add_item(items, rec)
    return items


def dedupe(items: list[dict]) -> list[dict]:
    by_key: dict[tuple[str, str], dict] = {}
    for item in items:
        key = (item["service_id"], item["configuration_item"].lower())
        if key not in by_key:
            by_key[key] = item
        else:
            current = by_key[key]
            for field in ["risk_reason", "security_value_reason", "references", "api_endpoint"]:
                if len(item.get(field, "")) > len(current.get(field, "")):
                    current[field] = item[field]
    return list(by_key.values())


def build_services(items: list[dict]) -> list[dict]:
    grouped = defaultdict(list)
    for item in items:
        grouped[item["service_id"]].append(item)
    for sid in SERVICE_META:
        grouped.setdefault(sid, [])
    services = []
    for sid, service_items in sorted(grouped.items(), key=lambda kv: SERVICE_META.get(kv[0], {}).get("service", kv[0])):
        meta = SERVICE_META.get(sid, {})
        service_name = meta.get("service", sid.replace("-", " ").title())
        focus = "; ".join(item["configuration_item"] for item in service_items[:4]) or meta.get("crypto_category", "N/A")
        risk_names = []
        for item in service_items:
            if item["risk_type"] not in risk_names:
                risk_names.append(item["risk_type"])
        typical = "; ".join(risk_names[:4]) or ("不可直接配置密码套件、TLS policy 或算法选择" if meta.get("configurable_status") == "不可直接配置" else "待补充")
        crypto_values = [item["risk_type"] for item in service_items[:3]]
        services.append({
            "service_id": sid,
            "service": service_name,
            "aws_category": meta.get("aws_category", "AWS Cloud Service"),
            "service_group": meta.get("service_group", "TLS / 安全策略"),
            "crypto_category": infer_crypto_category(sid, meta.get("crypto_category", "")) or "; ".join(crypto_values),
            "configurable_status": meta.get("configurable_status", "可配置"),
            "main_focus": meta.get("main_focus", focus) if meta.get("main_focus") else focus,
            "typical_risks": meta.get("typical_risks", typical) if meta.get("typical_risks") else typical,
            "service_intro": meta.get("service_intro", f"{service_name} 的密码相关配置已整理为服务卡片和配置项分析。"),
            "detail_path": f"#/services/{sid}",
        })
    return services


def build_risk_types() -> list[dict]:
    return [
        {
            "risk_id": risk_id,
            "risk_type": name,
            "risk_description": desc,
            "related_services": related,
            "typical_risky_values": risky,
            "recommended_direction": recommend,
            "notes": "依据 AWS 官方文档、密码协议安全实践和配置项安全影响整理。",
        }
        for risk_id, name, desc, related, risky, recommend in RISK_TYPES
    ]


def main():
    items: list[dict] = []
    summary = ROOT / "张聪艺-工作日报0430汇总表格.xlsx"
    if summary.exists():
        items.extend(extract_unified_summary(summary))
    for path in ROOT.glob("*.xlsx"):
        items.extend(extract_special_workbook(path))
    items = dedupe(items)
    services = build_services(items)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    (DATA_DIR / "config_items.json").write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "services.json").write_text(json.dumps(services, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "risk_types.json").write_text(json.dumps(build_risk_types(), ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"services={len(services)} config_items={len(items)} risk_types={len(RISK_TYPES)}")


if __name__ == "__main__":
    main()
