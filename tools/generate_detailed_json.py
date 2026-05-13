from __future__ import annotations

import json
import re
from collections import OrderedDict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "aws-risk-platform" / "data"


def text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def split_values(value: str) -> list[str]:
    return [part.strip() for part in re.split(r";|；|、|\n| / ", value or "") if part.strip()]


def slug(value: str, max_len: int = 58) -> str:
    value = value.lower().replace("&", " and ")
    value = re.sub(r"aws|amazon", "", value)
    value = value.replace("+", " plus ")
    value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
    return (value[:max_len].strip("-") or "item")


SERVICE_META = OrderedDict(
    [
        (
            "cloudfront",
            {
                "service": "Amazon CloudFront",
                "aws_category": "Networking & Content Delivery",
                "service_group": "TLS / 安全策略",
                "crypto_category": "TLS/HTTPS; Certificate",
                "configurable_status": "可配置",
                "main_focus": "ViewerCertificate.MinimumProtocolVersion; viewer/origin HTTPS; certificate attachment",
                "typical_risks": "允许 SSLv3/TLS 1.0/1.1；使用旧 CloudFront security policy；证书来源、算法或续期管理不当",
                "service_intro": "CloudFront 在 viewer 侧通过 security policy 控制最低 TLS 版本与可用 cipher，并通过 ACM/IAM 证书服务自定义域名 HTTPS。重点应检查 MinimumProtocolVersion、ViewerCertificate 和 origin HTTPS 配置。",
            },
        ),
        (
            "alb",
            {
                "service": "Elastic Load Balancing - Application Load Balancer (ALB)",
                "aws_category": "Networking & Content Delivery",
                "service_group": "TLS / 安全策略",
                "crypto_category": "TLS/HTTPS; Certificate",
                "configurable_status": "可配置",
                "main_focus": "CreateListener.SslPolicy / ModifyListener.SslPolicy; listener certificates; mutual TLS",
                "typical_risks": "使用 ELBSecurityPolicy-2016-08 等旧策略；启用 TLS 1.0/1.1、CBC、非前向保密套件；证书治理不当",
                "service_intro": "ALB 的 HTTPS listener 可直接配置 SSL policy 和证书，是 TLS 协议、cipher suite 与证书基线的重要控制点。现代基线应优先 TLS 1.2/1.3、FS/Res/PQ/FIPS 策略并按业务兼容性回退。",
            },
        ),
        (
            "nlb",
            {
                "service": "Elastic Load Balancing - Network Load Balancer (NLB)",
                "aws_category": "Networking & Content Delivery",
                "service_group": "TLS / 安全策略",
                "crypto_category": "TLS/HTTPS; Certificate",
                "configurable_status": "可配置",
                "main_focus": "TLS listener security policy; listener certificate",
                "typical_risks": "CLI/CloudFormation 未指定 policy 时落到旧默认策略；TLS 1.0/1.1 或非 FS cipher 被允许",
                "service_intro": "NLB TLS listener 通过 security policy 决定客户端协商协议和 cipher。AWS 文档显示控制台默认已偏向 2025 PQ/Res 策略，但非控制台创建仍可能默认到 ELBSecurityPolicy-2016-08。",
            },
        ),
        (
            "vpc-lattice",
            {
                "service": "Amazon VPC Lattice",
                "aws_category": "Networking & Content Delivery",
                "service_group": "TLS / 安全策略",
                "crypto_category": "TLS/HTTPS; Certificate",
                "configurable_status": "部分可配置",
                "main_focus": "listener/custom domain/certificate related settings",
                "typical_risks": "服务监听、域名和证书配置不一致；未使用 HTTPS 或 ACM 证书治理薄弱",
                "service_intro": "VPC Lattice 的密码相关面主要集中在 listener、custom domain 和证书绑定。它不像 ELB 暴露完整 cipher policy，但仍需要检查 HTTPS、证书和域名治理。",
            },
        ),
        (
            "route-53",
            {
                "service": "Amazon Route 53",
                "aws_category": "Networking & Content Delivery",
                "service_group": "DNSSEC / 完整性",
                "crypto_category": "DNSSEC",
                "configurable_status": "可配置",
                "main_focus": "EnableHostedZoneDNSSEC / CreateKeySigningKey",
                "typical_risks": "安全敏感 public hosted zone 未启用 DNSSEC；KSK/KMS 权限配置不当",
                "service_intro": "Route 53 的主要密码学配置是 DNSSEC signing。它通过 KSK 与 KMS 配合提供 DNS 数据来源验证，适用于需要防 DNS 篡改的 public hosted zone。",
            },
        ),
        (
            "vpn",
            {
                "service": "AWS Site-to-Site VPN",
                "aws_category": "Networking & Content Delivery",
                "service_group": "VPN / IPsec",
                "crypto_category": "IKE/IPsec; Encryption; Integrity; DH",
                "configurable_status": "可配置",
                "main_focus": "ModifyVpnConnectionOptions; Phase1/Phase2 algorithms; IKE version",
                "typical_risks": "IKEv1、弱 DH group、SHA1/低强度完整性算法、旧兼容加密套件",
                "service_intro": "Site-to-Site VPN 的密码配置集中在 IKE 版本、Phase 1/2 加密算法、完整性算法和 DH group。应按对端兼容性选择现代 IKEv2 与 AES-GCM/AES/SHA2/DH 高强度组合。",
            },
        ),
        (
            "kms",
            {
                "service": "AWS KMS",
                "aws_category": "Security, Identity, & Compliance",
                "service_group": "密钥管理",
                "crypto_category": "KMS; KeySpec; KeyUsage; Rotation; Origin",
                "configurable_status": "可配置",
                "main_focus": "CreateKey.KeySpec / KeyUsage / Origin / MultiRegion; Rotation; RSA/HMAC/ML-DSA algorithms",
                "typical_risks": "KeySpec/KeyUsage 与用途不匹配；RSAES_OAEP_SHA_1、HMAC_224 等兼容或低强度选择；外部密钥材料生命周期失控",
                "service_intro": "KMS 是 AWS 密钥管理核心。KeySpec 创建后不可变，决定密钥类型及可用算法；KeyUsage 决定加密、签名、MAC 或密钥协商用途。应避免用途混用并建立轮换、权限和外部密钥材料流程。",
            },
        ),
        (
            "cloudhsm",
            {
                "service": "AWS CloudHSM",
                "aws_category": "Security, Identity, & Compliance",
                "service_group": "HSM / 密码机制",
                "crypto_category": "HSM; PKCS#11; JCE; CNG",
                "configurable_status": "部分可配置",
                "main_focus": "PKCS#11/JCE/CNG mechanisms; key attributes; FIPS mode",
                "typical_risks": "使用 SHA1、DES3、ECB、RSA PKCS#1 v1.5 等旧机制；key attribute 权限过宽；Non-FIPS mode 不满足合规要求",
                "service_intro": "CloudHSM 暴露底层 HSM 机制和 key attribute，适合需要自管 HSM、PKCS#11/JCE/CNG 集成和高控制度的场景。风险多来自机制选择、权限属性和合规模式。",
            },
        ),
        (
            "acm",
            {
                "service": "AWS Certificate Manager",
                "aws_category": "Security, Identity, & Compliance",
                "service_group": "证书管理",
                "crypto_category": "Certificate; TLS",
                "configurable_status": "可配置",
                "main_focus": "RequestCertificate.KeyAlgorithm; ImportCertificate; validation; export/transparency options",
                "typical_risks": "RSA_1024 或目标服务不支持的算法；导入证书私钥和续期自管风险；证书导出扩大私钥暴露面",
                "service_intro": "ACM 负责 SSL/TLS X.509 证书申请、导入、存储、续期和部署。新申请证书当前支持 RSA_2048、EC_prime256v1、EC_secp384r1；导入证书算法范围更广但生命周期风险更高。",
            },
        ),
        (
            "signer",
            {
                "service": "AWS Signer",
                "aws_category": "Security, Identity, & Compliance",
                "service_group": "数字签名",
                "crypto_category": "Digital Signature; Code Signing",
                "configurable_status": "部分可配置",
                "main_focus": "PutSigningProfile / StartSigningJob; platformId; signing material",
                "typical_risks": "选择 SHA1/RSA 等旧平台；签名证书来源不清；签名权限过宽或未在部署链路强制校验",
                "service_intro": "AWS Signer 用于代码和制品签名。核心配置是 signing profile、platformId、签名材料和签名作业权限，应优先 SHA256/SHA384/ECDSA 等现代平台并在发布链路验证签名。",
            },
        ),
        (
            "encryption-sdk",
            {
                "service": "AWS Encryption SDK",
                "aws_category": "Developer Tools",
                "service_group": "客户端加密 SDK",
                "crypto_category": "Client-side Encryption; Keyring; Commitment Policy",
                "configurable_status": "可配置",
                "main_focus": "keyring; wrapping keys; commitment policy; algorithm suite; encryption context",
                "typical_risks": "裸 discovery keyring；raw keyring 管理不当；未要求 key commitment；encryption context 放敏感信息或不一致",
                "service_intro": "AWS Encryption SDK 提供信封加密和 keyring 抽象。应优先受控 KMS keyring、严格 key 标识、现代 commitment policy 和非敏感 encryption context。",
            },
        ),
        (
            "database-encryption-sdk",
            {
                "service": "AWS Database Encryption SDK",
                "aws_category": "Developer Tools",
                "service_group": "客户端加密 SDK",
                "crypto_category": "Client-side Encryption; Searchable Encryption; Keyring",
                "configurable_status": "可配置",
                "main_focus": "table encryption config; cryptographic actions; beacon configuration; keyring",
                "typical_risks": "敏感字段 DO_NOTHING；discovery mode 无 filter；beacon 泄露搜索模式；raw key 管理薄弱",
                "service_intro": "Database Encryption SDK 将字段级加密、签名和可搜索加密配置放在应用侧。重点检查 field actions、beacon、keyring 和 branch key/cache。",
            },
        ),
        (
            "s3-encryption-client",
            {
                "service": "Amazon S3 Encryption Client",
                "aws_category": "Storage",
                "service_group": "客户端加密 SDK",
                "crypto_category": "Client-side Encryption; Keyring; Wrapping Key",
                "configurable_status": "可配置",
                "main_focus": "wrapping key; keyring; encryption context; commitment policy",
                "typical_risks": "Raw AES/RSA wrapping key 自管不当；自定义 keyring 缺少审计；encryption context 明文泄露或字符集不兼容",
                "service_intro": "S3 Encryption Client 在客户端加密对象并把加密后的对象上传到 S3。它使用唯一 data key 加密对象，并用 wrapping key 保护 data key；没有成熟密钥管理时应优先 AWS KMS。",
            },
        ),
        (
            "payment-cryptography",
            {
                "service": "AWS Payment Cryptography",
                "aws_category": "Financial Services",
                "service_group": "支付密码",
                "crypto_category": "Payment Cryptography; PIN; CVV; TR-31; TR-34; HSM",
                "configurable_status": "可配置",
                "main_focus": "CreateKey.KeyAttributes; KeyAlgorithm; KeyUsage; KeyModesOfUse; export/import",
                "typical_risks": "TDES_2KEY 用于新系统；KeyUsage/KeyModesOfUse 过宽；exportable key 无业务必要；PIN/CVV/MAC/data encryption 用途混用",
                "service_intro": "AWS Payment Cryptography 用于支付行业密钥和密码操作。CreateKey 时 KeyClass、KeyAlgorithm、KeyUsage、KeyModesOfUse 等属性创建后不可变，应严格按 TR-31/TR-34 和业务用途约束。",
            },
        ),
        (
            "secrets-manager",
            {
                "service": "AWS Secrets Manager",
                "aws_category": "Security, Identity, & Compliance",
                "service_group": "密钥管理",
                "crypto_category": "Secrets encryption; KMS; Rotation",
                "configurable_status": "可配置",
                "main_focus": "KMS key selection; rotation schedule; resource policy",
                "typical_risks": "使用默认 key 不符合隔离要求；未轮换长期凭证；resource policy 过宽",
                "service_intro": "Secrets Manager 本身不是密码算法配置服务，但 secret 静态加密依赖 KMS，并承担凭证轮换和访问控制。适合在风险平台中作为密钥/凭证生命周期控制项。",
            },
        ),
        (
            "private-ca",
            {
                "service": "AWS Private Certificate Authority",
                "aws_category": "Security, Identity, & Compliance",
                "service_group": "证书管理",
                "crypto_category": "Private PKI; Certificate; CA signing algorithm",
                "configurable_status": "可配置",
                "main_focus": "CA key algorithm; signing algorithm; certificate templates; revocation",
                "typical_risks": "CA 算法与签发证书族不匹配；吊销与审计配置不足；私有 CA 权限过宽",
                "service_intro": "AWS Private CA 为私有 PKI 和 ACM 私有证书提供 CA 能力。它与 ACM 的证书请求算法、CA signing algorithm 和生命周期治理密切相关。",
            },
        ),
        (
            "vpc",
            {
                "service": "Amazon VPC",
                "aws_category": "Networking & Content Delivery",
                "service_group": "不可直接配置",
                "crypto_category": "N/A",
                "configurable_status": "不可直接配置",
                "main_focus": "N/A; rely on attached services such as VPN, PrivateLink, TLS endpoints",
                "typical_risks": "误把网络隔离替代传输层加密；未在上层服务配置 TLS、KMS 或证书",
                "service_intro": "VPC 本身不提供 TLS/cipher 选择项，密码风险通常来自挂载的 VPN、负载均衡、PrivateLink、终端服务或应用层 TLS 配置。",
            },
        ),
        (
            "privatelink",
            {
                "service": "AWS PrivateLink",
                "aws_category": "Networking & Content Delivery",
                "service_group": "不可直接配置",
                "crypto_category": "N/A",
                "configurable_status": "不可直接配置",
                "main_focus": "N/A; service endpoint and consumer/provider access controls",
                "typical_risks": "把私网连接等同于端到端加密；endpoint policy/服务权限过宽",
                "service_intro": "PrivateLink 主要提供私网连通性，不直接暴露密码套件配置。应结合服务端 HTTPS/TLS、证书和访问控制检查。",
            },
        ),
        (
            "cloud-map",
            {
                "service": "AWS Cloud Map",
                "aws_category": "Networking & Content Delivery",
                "service_group": "不可直接配置",
                "crypto_category": "N/A",
                "configurable_status": "不可直接配置",
                "main_focus": "N/A; service discovery records",
                "typical_risks": "服务发现配置不等于传输加密；实际通信仍需依赖应用或负载均衡 TLS",
                "service_intro": "Cloud Map 负责服务发现，不直接配置密码算法、TLS policy 或证书。风险平台中应标为非直接配置项，并提示检查被发现服务自身的 TLS/KMS 设置。",
            },
        ),
    ]
)


RISK_TYPES = [
    ("old-tls", "旧 TLS/SSL 协议风险", "允许 SSLv3、TLS 1.0 或 TLS 1.1 会带来协议降级、弱算法和合规风险。RFC 8996 已明确弃用 TLS 1.0/1.1。", "CloudFront; ALB; NLB; Site-to-Site VPN", "SSLv3; TLSv1; TLSv1.1; legacy security policies", "优先 TLS 1.2/1.3；新建入口使用 AWS 现代 security policy；VPN 优先 IKEv2 与现代 Phase 1/2 组合"),
    ("old-security-policy", "旧 TLS 安全策略/密码套件风险", "ELB、CloudFront 等 security policy 若包含 CBC、非 FS、SHA1 或旧协议，会降低传输安全性。", "CloudFront; ALB; NLB", "ELBSecurityPolicy-2016-08; TLSv1_2016; policies supporting CBC/non-FS ciphers", "优先 TLS13/TLS12 Res、PQ、FIPS 或 FS 策略；按客户端兼容性最小化回退"),
    ("certificate", "证书与私钥生命周期风险", "证书算法、来源、导入私钥、导出和续期策略会影响 TLS 身份验证和私钥暴露面。", "ACM; Private CA; CloudFront; ALB; NLB; Signer", "RSA_1024; expired imported certificate; export enabled without need; unmanaged third-party certificate", "优先 ACM 托管证书和 DNS 验证；导入证书需建立私钥保护、链验证和续期流程"),
    ("dnssec", "DNSSEC 配置风险", "安全敏感 public hosted zone 未启用 DNSSEC 会增加 DNS 篡改与缓存投毒风险。", "Route 53", "DNSSEC signing disabled; KSK/KMS permission broken", "为需要完整性保护的 public hosted zone 启用 DNSSEC signing，并维护 KSK 与 KMS 权限"),
    ("kms-key", "KMS 密钥规格/用途风险", "KeySpec、KeyUsage、Origin 和算法选择创建后难以更改，错误设计会导致用途混用、强度不足或数据不可恢复。", "AWS KMS", "HMAC_224; RSAES_OAEP_SHA_1; mismatched KeyUsage; unmanaged EXTERNAL key material", "按用途选择 KeySpec/KeyUsage；优先对称 KMS key、RSA_OAEP_SHA_256、HMAC_256+；管理外部密钥材料生命周期"),
    ("key-lifecycle", "密钥与证书生命周期风险", "未轮换、未监控过期、外部密钥材料或导入证书缺少重导入流程，会让长期凭证和证书成为系统性风险。", "KMS; ACM; Secrets Manager; Payment Cryptography", "rotation disabled without justification; expired certificate; missing imported key material expiration/reimport", "为 customer managed key、secret、证书和支付密钥建立轮换、过期告警和再导入流程"),
    ("hsm-mechanism", "HSM 机制与 key attribute 风险", "CloudHSM 允许底层机制和 key attribute 配置，弱机制或权限过宽会造成签名、解密、导出等能力滥用。", "CloudHSM", "SHA1; DES3; AES-ECB; RSA PKCS#1 v1.5; extractable=true without need", "优先 AES-GCM、RSA-OAEP/PSS、SHA256+；最小化 key attribute 权限，合规场景使用 FIPS mode"),
    ("code-signing", "代码签名治理风险", "代码签名平台、证书、哈希/签名算法和权限若不受控，会削弱制品完整性保证。", "AWS Signer", "SHA1 platform; untrusted signing certificate; excessive StartSigningJob permission", "选择现代签名平台，使用受控证书并在部署/运行时强制验证签名"),
    ("sdk-policy", "客户端加密 SDK 策略风险", "keyring、wrapping key、commitment policy 和 algorithm suite 配置不当会扩大解密范围或降低密文完整性保证。", "AWS Encryption SDK; Database Encryption SDK; S3 Encryption Client", "裸 discovery keyring; raw key unmanaged; commitment disabled; overly broad KMS key IDs", "优先 KMS keyring、严格 ARN、discovery filter、require encrypt/decrypt commitment"),
    ("encryption-context", "Encryption context/AAD 风险", "encryption context 是非秘密 AAD，常被记录或明文存储；放入敏感信息或加解密不一致会造成泄露或不可用。", "AWS Encryption SDK; Database Encryption SDK; S3 Encryption Client; KMS", "包含敏感数据; 非 US-ASCII 导致兼容问题; 加解密 context 不一致", "仅放非敏感、稳定、可审计的键值；S3 Encryption Client 场景优先 US-ASCII 并保持加解密一致"),
    ("searchable-encryption", "可搜索加密/beacon 风险", "beacon 会泄露一定搜索模式和频率信息，低基数或敏感字段配置不当会降低隐私保护。", "Database Encryption SDK", "低基数字段 beacon; 过长/过短 beacon; 未做威胁建模", "仅对必要字段启用 beacon，按查询需求、基数和威胁模型选择 beacon 长度与类型"),
    ("payment-key", "支付密钥属性与用途风险", "支付密钥属性创建后不可变，KeyClass、KeyAlgorithm、KeyUsage、KeyModesOfUse 不匹配会造成合规或操作风险。", "AWS Payment Cryptography", "TDES_2KEY for new systems; exportable without need; mixed PIN/CVV/MAC/data usage", "按 TR-31/TR-34 与业务用途创建最小权限 key；新系统优先 AES/RSA/ECC 合规强度组合"),
    ("not-configurable", "无直接密码配置但易误判", "VPC、PrivateLink、Cloud Map 等服务本身不提供 cipher/TLS policy 配置，不能替代应用层或关联服务的加密控制。", "VPC; PrivateLink; Cloud Map", "N/A treated as safe control; missing TLS on attached service", "标明无直接配置项，并回溯实际承载通信的 ALB/NLB/API/应用 TLS 与 KMS 配置"),
    ("access-policy", "访问控制扩大密码能力风险", "密钥、证书、签名和加密 SDK 的安全不仅取决于算法，还取决于谁能解密、签名、导出或修改配置。", "KMS; CloudHSM; Signer; Payment Cryptography; Secrets Manager", "wildcard kms:Decrypt; broad signer:StartSigningJob; extract/export permissions without approval", "使用最小权限、条件键、职责分离、审计和变更审批约束高风险密码操作"),
]


RISK_BY_SERVICE = {
    "cloudfront": "旧 TLS 安全策略/密码套件风险",
    "alb": "旧 TLS 安全策略/密码套件风险",
    "nlb": "旧 TLS 安全策略/密码套件风险",
    "vpc-lattice": "证书与私钥生命周期风险",
    "route-53": "DNSSEC 配置风险",
    "vpn": "旧 TLS/SSL 协议风险",
    "kms": "KMS 密钥规格/用途风险",
    "cloudhsm": "HSM 机制与 key attribute 风险",
    "acm": "证书与私钥生命周期风险",
    "private-ca": "证书与私钥生命周期风险",
    "signer": "代码签名治理风险",
    "encryption-sdk": "客户端加密 SDK 策略风险",
    "database-encryption-sdk": "客户端加密 SDK 策略风险",
    "s3-encryption-client": "客户端加密 SDK 策略风险",
    "payment-cryptography": "支付密钥属性与用途风险",
    "secrets-manager": "密钥与证书生命周期风险",
    "vpc": "无直接密码配置但易误判",
    "privatelink": "无直接密码配置但易误判",
    "cloud-map": "无直接密码配置但易误判",
}


REFS = {
    "cloudfront": "https://docs.aws.amazon.com/AmazonCloudFront/latest/DeveloperGuide/secure-connections-supported-viewer-protocols-ciphers.html ; https://docs.aws.amazon.com/cloudfront/latest/APIReference/API_ViewerCertificate.html",
    "alb": "https://docs.aws.amazon.com/elasticloadbalancing/latest/application/describe-ssl-policies.html ; https://docs.aws.amazon.com/cli/latest/reference/elbv2/modify-listener.html",
    "nlb": "https://docs.aws.amazon.com/elasticloadbalancing/latest/network/describe-ssl-policies.html ; https://docs.aws.amazon.com/cli/latest/reference/elbv2/modify-listener.html",
    "vpc-lattice": "https://docs.aws.amazon.com/vpc-lattice/latest/ug/what-is-vpc-lattice.html ; https://docs.aws.amazon.com/vpc-lattice/latest/ug/data-protection.html",
    "route-53": "https://docs.aws.amazon.com/Route53/latest/DeveloperGuide/dns-configuring-dnssec.html ; https://docs.aws.amazon.com/cli/latest/reference/route53/enable-hosted-zone-dnssec.html",
    "vpn": "https://docs.aws.amazon.com/cli/latest/reference/ec2/modify-vpn-connection-options.html ; https://docs.aws.amazon.com/vpn/latest/s2svpn/security.html",
    "kms": "https://docs.aws.amazon.com/kms/latest/developerguide/asymmetric-key-specs.html ; https://docs.aws.amazon.com/kms/latest/developerguide/mldsa.html",
    "cloudhsm": "https://docs.aws.amazon.com/cloudhsm/latest/userguide/pkcs11-mechanisms.html",
    "acm": "https://docs.aws.amazon.com/acm/latest/APIReference/API_RequestCertificate.html ; https://docs.aws.amazon.com/cli/latest/reference/acm/request-certificate.html ; https://docs.aws.amazon.com/acm/latest/APIReference/API_ImportCertificate.html",
    "private-ca": "https://docs.aws.amazon.com/privateca/latest/userguide/PcaWelcome.html",
    "signer": "https://docs.aws.amazon.com/signer/latest/api/API_PutSigningProfile.html ; https://docs.aws.amazon.com/signer/latest/api/API_StartSigningJob.html",
    "encryption-sdk": "https://docs.aws.amazon.com/encryption-sdk/latest/developer-guide/concepts.html ; https://docs.aws.amazon.com/encryption-sdk/latest/developer-guide/concepts.html#keyrings",
    "database-encryption-sdk": "https://docs.aws.amazon.com/database-encryption-sdk/latest/devguide/what-is-database-encryption-sdk.html ; https://docs.aws.amazon.com/database-encryption-sdk/latest/devguide/configure.html",
    "s3-encryption-client": "https://docs.aws.amazon.com/amazon-s3-encryption-client/latest/developerguide/concepts.html ; https://docs.aws.amazon.com/AmazonS3/latest/userguide/UsingClientSideEncryption.html",
    "payment-cryptography": "https://docs.aws.amazon.com/payment-cryptography/latest/APIReference/API_CreateKey.html ; https://docs.aws.amazon.com/payment-cryptography/latest/APIReference/API_KeyAttributes.html ; https://docs.aws.amazon.com/payment-cryptography/latest/userguide/keys-validattributes.html",
    "secrets-manager": "https://docs.aws.amazon.com/secretsmanager/latest/userguide/security-encryption.html ; https://docs.aws.amazon.com/secretsmanager/latest/userguide/rotating-secrets.html",
    "vpc": "https://docs.aws.amazon.com/vpc/latest/userguide/what-is-amazon-vpc.html",
    "privatelink": "https://docs.aws.amazon.com/vpc/latest/privatelink/what-is-privatelink.html",
    "cloud-map": "https://docs.aws.amazon.com/cloud-map/latest/dg/what-is-cloud-map.html",
}


def service_id_from_name(name: str) -> str | None:
    n = name.lower().replace("_", " ")
    if "payment" in n:
        return "payment-cryptography"
    if "database encryption sdk" in n:
        return "database-encryption-sdk"
    if "s3 encryption client" in n:
        return "s3-encryption-client"
    if "encryption sdk" in n and "database" not in n:
        return "encryption-sdk"
    if "signer" in n:
        return "signer"
    if "acm" in n:
        return "acm"
    if "kms" in n:
        return "kms"
    if "hsm" in n:
        return "cloudhsm"
    if "nlb" in n:
        return "nlb"
    if "alb" in n:
        return "alb"
    if "vpn" in n:
        return "vpn"
    return None


def service_id_from_service_cell(value: str) -> str:
    n = value.lower().replace("_", " ")
    aliases = [
        ("cloudfront", "cloudfront"),
        ("application load balancer", "alb"),
        ("alb", "alb"),
        ("network load balancer", "nlb"),
        ("nlb", "nlb"),
        ("vpc lattice", "vpc-lattice"),
        ("route 53", "route-53"),
        ("site-to-site vpn", "vpn"),
        ("virtual private network", "vpn"),
        ("kms", "kms"),
        ("cloudhsm", "cloudhsm"),
        ("certificate manager", "acm"),
        ("acm", "acm"),
        ("private ca", "private-ca"),
        ("signer", "signer"),
        ("database encryption sdk", "database-encryption-sdk"),
        ("s3 encryption client", "s3-encryption-client"),
        ("encryption sdk", "encryption-sdk"),
        ("payment cryptography", "payment-cryptography"),
        ("secrets manager", "secrets-manager"),
        ("privatelink", "privatelink"),
        ("private link", "privatelink"),
        ("cloud map", "cloud-map"),
        ("vpc", "vpc"),
    ]
    for needle, sid in aliases:
        if needle in n:
            return sid
    return slug(value, 32)


def find_header(rows: list[list[str]], required: list[str]) -> tuple[int, list[str]] | None:
    for idx, row in enumerate(rows[:30]):
        joined = " ".join(row).lower()
        if all(req.lower() in joined for req in required):
            return idx, row
    return None


def header_map(headers: list[str]) -> dict[str, int]:
    result = {}
    for idx, header in enumerate(headers):
        h = header.lower()
        if not h:
            continue
        if "service" == h or "service" in h and "public" not in h:
            result.setdefault("service", idx)
        if "description" in h or "说明" in h:
            result.setdefault("description", idx)
        if "crypto_category" in h:
            result.setdefault("crypto_category", idx)
        if "configuration_item" in h or "配置项" in h or "页面列出的项目" in h:
            result.setdefault("configuration_item", idx)
        if "api_endpoint" in h or "对应命令" in h or "api" in h or "cli" in h:
            result.setdefault("api_endpoint", idx)
        if "recommended" in h or "安全值" in h or "推荐值" in h:
            result.setdefault("recommended_values", idx)
        if "risky" in h or "风险值" in h or "注意值" in h:
            result.setdefault("risky_values", idx)
        if "references" in h or "参考" in h or "依据" in h:
            result.setdefault("references", idx)
        if "原因" in h or "典型风险点" in h:
            result.setdefault("risk_reason", idx)
        if "可选值" in h or "典型值" in h or "表现" in h:
            result.setdefault("possible_values", idx)
    return result


def cell(row: list[str], mapping: dict[str, int], key: str) -> str:
    idx = mapping.get(key)
    return text(row[idx]) if idx is not None and idx < len(row) else ""


def risk_level(risky: str, reason: str) -> str:
    s = f"{risky} {reason}".lower()
    high_tokens = ["sslv3", "tlsv1", "tls 1.0", "tls 1.1", "rsa_1024", "sha1", "tdes_2key", "do_nothing", "裸", "过宽", "disabled", "exportable"]
    medium_tokens = ["rsa_2048", "sha224", "cbc", "pkcs", "raw", "discovery", "non-fips", "兼容", "注意"]
    if any(token in s for token in high_tokens):
        return "高"
    if any(token in s for token in medium_tokens):
        return "中"
    return "中"


def build_reason(service_id: str, config: str, risky: str, reason: str) -> str:
    if reason:
        return reason
    risk = RISK_BY_SERVICE.get(service_id, "密码配置风险")
    if risky:
        return f"{config} 若取值为 {risky}，会触发“{risk}”：可能降低算法强度、扩大密钥/证书暴露面，或造成合规与兼容性风险。"
    return f"{config} 是 {SERVICE_META.get(service_id, {}).get('service', service_id)} 的密码/加密相关控制点，需要结合业务场景确认强度、用途和权限边界。"


def build_security_reason(service_id: str, config: str, recommended: str) -> str:
    if recommended:
        return f"{recommended} 通常更符合现代密码基线、最小权限或托管生命周期要求；仍需按客户端兼容性、合规框架和业务连续性验证。"
    return f"建议以 AWS 官方当前支持值为准，优先选择现代算法、托管证书/密钥生命周期和严格权限边界。"


def add_item(items: OrderedDict, service_id: str, config: str, api: str, recommended: str, risky: str, references: str = "", reason: str = ""):
    config = text(config)
    if not config or len(config) < 2:
        return
    if config in {"配置项", "Configuration_Item", "页面列出的项目"}:
        return
    service_id = service_id if service_id in SERVICE_META else service_id_from_service_cell(service_id)
    item_id = f"{service_id}_{slug(config)}"
    if item_id in items:
        existing = items[item_id]
        for key, value in {
            "api_endpoint": api,
            "recommended_values": recommended,
            "risky_values": risky,
            "references": references,
        }.items():
            value = text(value)
            if value and value not in existing[key]:
                existing[key] = f"{existing[key]}; {value}" if existing[key] else value
        return
    items[item_id] = {
        "item_id": item_id,
        "service_id": service_id,
        "configuration_item": config,
        "api_endpoint": text(api) or config,
        "recommended_values": text(recommended) or "按业务场景选择 AWS 当前支持的现代安全值，并记录例外原因",
        "risky_values": text(risky) or "未配置、配置过宽、使用旧兼容值或与业务用途不匹配",
        "risk_type": RISK_BY_SERVICE.get(service_id, "密码配置风险"),
        "risk_level": risk_level(risky, reason),
        "risk_reason": build_reason(service_id, config, risky, reason),
        "security_value_reason": build_security_reason(service_id, config, recommended),
        "references": text(references) or REFS.get(service_id, ""),
    }


def rows_from_sheet(ws) -> list[list[str]]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = [text(v) for v in row]
        if any(values):
            rows.append(values)
    return rows


def extract_generic_tables(items: OrderedDict):
    for path in sorted(ROOT.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        if service_id_from_name(path.name) in {"alb", "nlb"}:
            # The ALB/NLB workbooks contain detailed policy matrices plus older
            # hand-built summary sheets. Use the matrices below instead; several
            # summary rows split tokens or invert VPN recommended/risky values.
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        workbook_sid = service_id_from_name(path.name)
        for ws in wb.worksheets:
            rows = rows_from_sheet(ws)
            if not rows:
                continue

            # 0430 consolidated English table.
            found = find_header(rows, ["Service"])
            if found and any("Recommended" in h or "Risky" in h for h in found[1]):
                idx, headers = found
                mapping = header_map(headers)
                last_sid = ""
                last_desc = ""
                for row in rows[idx + 1 :]:
                    sid = service_id_from_service_cell(cell(row, mapping, "service") or last_sid)
                    if cell(row, mapping, "service"):
                        last_sid = cell(row, mapping, "service")
                    if cell(row, mapping, "description"):
                        last_desc = cell(row, mapping, "description")
                    config = cell(row, mapping, "configuration_item") or cell(row, mapping, "api_endpoint") or last_desc
                    add_item(
                        items,
                        sid,
                        config,
                        cell(row, mapping, "api_endpoint"),
                        cell(row, mapping, "recommended_values"),
                        cell(row, mapping, "risky_values"),
                        cell(row, mapping, "references"),
                        cell(row, mapping, "risk_reason"),
                    )
                continue

            found = find_header(rows, ["配置项"])
            if not found:
                continue
            idx, headers = found
            mapping = header_map(headers)
            sid = workbook_sid
            sheet_name = ws.title.lower()
            if "cloudhsm" in sheet_name or "hsm" in sheet_name or ws.title in {"0421", "key attributes", "整体"}:
                sid = "cloudhsm"
            elif "kms" in sheet_name or ws.title in {"总体表"}:
                sid = "kms"
            if sid is None:
                continue

            for row in rows[idx + 1 :]:
                config = cell(row, mapping, "configuration_item")
                possible = cell(row, mapping, "possible_values")
                recommended = cell(row, mapping, "recommended_values")
                risky = cell(row, mapping, "risky_values")
                if possible and possible not in recommended and possible not in risky:
                    config = f"{config}（可选值：{possible}）" if config and len(possible) < 180 else config
                add_item(
                    items,
                    sid,
                    config,
                    cell(row, mapping, "api_endpoint"),
                    recommended,
                    risky,
                    cell(row, mapping, "references"),
                    cell(row, mapping, "risk_reason"),
                )
        wb.close()


def extract_tls_policies(items: OrderedDict):
    for path in sorted(ROOT.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        sid = service_id_from_name(path.name)
        if sid not in {"alb", "nlb"}:
            continue
        # Keep one ALB source to avoid exact duplicate 0409/0410 workbooks.
        if sid == "alb" and "0410" not in path.name:
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows = rows_from_sheet(ws)
            found = find_header(rows, ["Security policies", "TLS 1.2"])
            if not found:
                continue
            idx, headers = found
            for row in rows[idx + 1 :]:
                policy = row[0] if row else ""
                if not policy.startswith("ELBSecurityPolicy"):
                    continue
                vals = {headers[i]: row[i] if i < len(row) else "" for i in range(min(len(headers), 5))}
                supports_old = vals.get("TLS 1.1") == "Yes" or vals.get("TLS 1.0") == "Yes"
                supports_13 = vals.get("TLS 1.3") == "Yes"
                supports_12 = vals.get("TLS 1.2") == "Yes"
                recommended = []
                risky = []
                if supports_13:
                    recommended.append("支持 TLS 1.3")
                if supports_12:
                    recommended.append("支持 TLS 1.2")
                if "FIPS" in policy:
                    recommended.append("FIPS policy，适合合规场景")
                if "PQ" in policy:
                    recommended.append("包含 PQ/后量子过渡方向策略")
                if "Res" in policy or "FS" in policy:
                    recommended.append("偏现代/前向保密策略")
                if supports_old:
                    risky.append("支持 TLS 1.0/1.1")
                if not supports_old and ("2016-08" not in policy):
                    risky.append("主要风险来自客户端兼容性回退，而非策略本身")
                add_item(
                    items,
                    sid,
                    f"TLS security policy: {policy}",
                    "CreateListener.SslPolicy / ModifyListener.SslPolicy / DescribeSslPolicies",
                    "; ".join(recommended) or "按兼容性选择 TLS 1.2/1.3 policy",
                    "; ".join(risky) or "避免选择支持 TLS 1.0/1.1、CBC 或非 FS cipher 的旧策略",
                    REFS[sid],
                    "TLS security policy 决定 listener 可协商的协议版本和 cipher suite；旧协议或旧 cipher 会扩大降级和合规风险。",
                )
        wb.close()


def extract_cloudhsm_mechanisms(items: OrderedDict):
    for path in sorted(ROOT.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        if "HSM" not in path.name and "0422" not in path.name:
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            if "机制" not in ws.title:
                continue
            for row in rows_from_sheet(ws):
                joined = " ".join(row)
                mechanisms = re.findall(r"\bCKM_[A-Z0-9_]+\b|(?<![A-Za-z])(?:sha1|sha224|sha256|sha384|sha512|ecdsa|rsa-pkcs-pss|rsa-pkcs|aes-gcm)(?![A-Za-z])", joined, re.I)
                for mech in mechanisms:
                    mech_u = mech.upper()
                    s = joined.lower()
                    risky = []
                    safe = []
                    if any(t in mech_u for t in ["SHA1", "SHA_1", "DES3", "ECB"]):
                        risky.append("旧/弱或不推荐机制，需仅限兼容场景")
                    if "PKCS" in mech_u and "OAEP" not in mech_u and "PSS" not in mech_u:
                        risky.append("RSA PKCS#1 v1.5 相关机制优先级低于 OAEP/PSS")
                    if any(t in mech_u for t in ["GCM", "OAEP", "PSS", "SHA256", "SHA384", "SHA512", "ECDSA"]):
                        safe.append("现代推荐机制或哈希强度较高")
                    add_item(
                        items,
                        "cloudhsm",
                        f"CloudHSM mechanism: {mech_u}",
                        "PKCS#11 / CloudHSM CLI mechanism selection",
                        "; ".join(safe) or "按合规和业务需求选择 AWS CloudHSM 当前支持机制，优先 AEAD、OAEP/PSS、SHA256+",
                        "; ".join(risky) or "机制与 key attribute 不匹配、兼容性回退或权限过宽",
                        REFS["cloudhsm"],
                        joined[:500],
                    )
        wb.close()


def add_manual_items(items: OrderedDict):
    manual = [
        ("cloudfront", "DistributionConfig.ViewerCertificate.MinimumProtocolVersion", "UpdateDistribution / ViewerCertificate.MinimumProtocolVersion", "TLSv1.3_2025; TLSv1.2_2025; TLSv1.2_2021", "SSLv3; TLSv1; TLSv1_2016; TLSv1.1_2016; TLSv1.2_2018/2019 仅作兼容回退"),
        ("cloudfront", "Viewer certificate for alternate domain names", "ViewerCertificate.ACMCertificateArn / IAMCertificateId / CloudFrontDefaultCertificate", "自定义域名优先 ACM 证书；证书算法和链满足目标客户端", "默认 CloudFront 证书用于自定义域名需求；过期、弱算法或无人管理的导入证书"),
        ("route-53", "EnableHostedZoneDNSSEC / CreateKeySigningKey", "EnableHostedZoneDNSSEC / CreateKeySigningKey", "为安全敏感 public hosted zone 启用 DNSSEC；KSK 使用受控 KMS 权限", "DNSSEC disabled；KSK inactive；KMS key policy 阻止 Route 53 使用"),
        ("vpn", "IKE version", "ModifyVpnConnectionOptions.IKEVersions", "IKEv2；仅在遗留对端强依赖时保留 IKEv1 兼容说明", "IKEv1 作为默认或长期配置；未记录兼容性例外"),
        ("vpn", "Phase 1 encryption algorithms", "ModifyVpnConnectionOptions.Phase1EncryptionAlgorithms", "AES128-GCM-16；AES256-GCM-16；AES128/AES256 结合 SHA2 完整性算法", "DES/3DES 或低强度兼容算法；只为旧设备保留的弱算法"),
        ("vpn", "Phase 2 encryption algorithms", "ModifyVpnConnectionOptions.Phase2EncryptionAlgorithms", "AES128-GCM-16；AES256-GCM-16；按对端支持选择 AES-GCM 优先", "DES/3DES；与对端协商导致长期落到弱算法"),
        ("vpn", "Phase 1/2 integrity algorithms", "ModifyVpnConnectionOptions.Phase1IntegrityAlgorithms / Phase2IntegrityAlgorithms", "SHA2-256、SHA2-384、SHA2-512；GCM 模式下按 AWS/对端要求配置", "SHA1 作为长期基线；完整性算法与加密模式或对端能力不匹配"),
        ("vpn", "Diffie-Hellman group numbers", "ModifyVpnConnectionOptions.Phase1DHGroupNumbers / Phase2DHGroupNumbers", "Group 14+，优先 19/20/21 等更高强度椭圆曲线组（取决于对端支持）", "Group 1/2/5 等低强度组；没有记录例外原因的旧组"),
        ("vpc-lattice", "TLS listener / custom domain certificate settings", "VPC Lattice listener/custom domain/certificate configuration", "HTTPS listener；使用 ACM 证书；域名和证书匹配", "HTTP 暴露敏感流量；证书来源不明；域名与证书不匹配"),
        ("secrets-manager", "Secret encryption KMS key", "CreateSecret.KmsKeyId / UpdateSecret.KmsKeyId", "按环境/应用隔离使用 customer managed KMS key；默认 key 仅用于低隔离要求场景", "跨环境共用 key；kms:Decrypt 权限过宽；未记录 key 选择依据"),
        ("secrets-manager", "Secret rotation", "RotateSecret / RotationRules", "数据库、API token 等长期 secret 建立自动或准自动轮换", "长期凭证不轮换；轮换 Lambda 权限过宽；应用无法平滑加载新 secret"),
        ("private-ca", "CA key algorithm and signing family", "CreateCertificateAuthority / IssueCertificate signing algorithm", "CA 与签发证书算法族匹配；按安全级别选择 RSA/ECDSA 强度", "CA signing algorithm 与证书请求不匹配；使用弱算法或无吊销策略"),
        ("vpc", "VPC network boundary is not encryption control", "N/A", "把 VPC 作为网络隔离控制，同时在上层服务配置 TLS/KMS/证书", "将私网误认为传输加密；敏感流量未启用 TLS"),
        ("privatelink", "PrivateLink endpoint service encryption dependency", "Endpoint service / endpoint policy", "使用 PrivateLink 限制网络路径，并在服务端启用 HTTPS/TLS 和访问控制", "只依赖 PrivateLink 不配置 TLS；endpoint policy 过宽"),
        ("cloud-map", "Service discovery records", "Cloud Map namespace/service/instance records", "服务发现只负责定位，实际服务使用 TLS/mTLS 或应用层加密", "把服务发现当作安全通道；发现到的服务端点未启用 TLS"),
    ]
    for sid, config, api, rec, risky in manual:
        add_item(items, sid, config, api, rec, risky, REFS.get(sid, ""))


def is_generic_value(value: str) -> bool:
    value = text(value)
    return (
        not value
        or value.startswith("按业务场景选择 AWS 当前支持的现代安全值")
        or value.startswith("未配置、配置过宽、使用旧兼容值")
    )


def merge_field(old: str, new: str, limit: int = 520) -> str:
    old = text(old)
    new = text(new)
    if not new:
        return old
    if not old or is_generic_value(old):
        return new
    if is_generic_value(new) or new in old:
        return old
    parts = []
    for part in re.split(r";|；", f"{old}; {new}"):
        part = text(part)
        if part and part not in parts:
            parts.append(part)
    merged = "; ".join(parts)
    return merged[:limit].rstrip(" ;")


def canonical_config(service_id: str, config: str) -> str | None:
    raw = text(config)
    lower = raw.lower()

    if not raw:
        return None
    if raw.startswith("TLS security policy:"):
        return raw
    if raw in {"安全", "Mode"}:
        return None
    if "（可选值：" in raw:
        raw = re.sub(r"（可选值：.*?）", "", raw).strip()
        lower = raw.lower()

    if service_id == "cloudfront":
        if "minimumprotocolversion" in lower or "minimum viewer tls" in lower:
            return "DistributionConfig.ViewerCertificate.MinimumProtocolVersion"
        if "viewer certificate" in lower:
            return "DistributionConfig.ViewerCertificate"

    if service_id == "acm":
        if "passphrase" in lower or "口令" in raw:
            return "ExportCertificate.Passphrase"
        if "keyalgorithm" in lower or "key algorithm" in lower or "证书密钥算法" in raw:
            return "RequestCertificate.KeyAlgorithm"
        if "importcertificate" in lower or "imported certificate" in lower or "导入证书" in raw:
            return "ImportCertificate / imported certificate lifecycle"
        if "validationmethod" in lower or "域名验证" in raw:
            return "ValidationMethod"
        if "certificatetransparency" in lower or "透明度日志" in raw or "ct 日志" in lower:
            return "CertificateTransparencyLoggingPreference"
        if "pinning" in lower or "固定" in raw:
            return "Certificate pinning"
        if "subject alternative" in lower or "san" in lower or "wildcard" in lower or "域名范围" in raw:
            return "Subject Alternative Names / wildcard"
        if "options.export" in lower or "证书导出" in raw or raw == "证书导出 Export":
            return "Options.Export"
        if "private ca" in lower or "certificateauthorityarn" in lower or "私有 ca" in raw or "私有证书" in raw:
            return "CertificateAuthorityArn / private CA"
        if "tls policy" in lower or "集成服务 tls" in lower:
            return "Integrated service TLS policy"
        if "renewal" in lower or "自动续期" in raw:
            return "Managed renewal"
        if raw in {"标签 Tags", "访问控制 / KMS grant 限制", "ACM 与集成服务 TLS 策略边界"}:
            return None

    if service_id == "encryption-sdk":
        if "commitment policy" in lower:
            return "Commitment policy"
        if "key commitment" in lower:
            return "Key commitment"
        if "discovery filter" in lower:
            return "Discovery filter"
        if "keyring" in lower:
            return "Keyring configuration"
        if "wrapping key" in lower or "generator key" in lower:
            return "Wrapping keys / KMS keys"
        if "kms key 类型" in raw or "kms key type" in lower:
            return "KMS key type"
        if "kms key 标识" in raw:
            return "KMS key identifier"
        if "encryption context" in lower:
            return "Encryption context"
        if "algorithm suite" in lower:
            return "Algorithm suite"
        if "digital signatures" in lower or "签名" in raw:
            return "Digital signatures"
        if "strict mode" in lower or "discovery mode" in lower:
            return "Strict mode / discovery mode"
        if raw == "是否使用 AWS KMS keyring":
            return "Keyring configuration"

    if service_id == "database-encryption-sdk":
        if "discovery filter" in lower:
            return "Discovery filter"
        if "strict" in lower or "discovery mode" in lower:
            return "Strict mode / discovery mode"
        if "keyring" in lower or "branch key cache" in lower:
            return "Keyring configuration"
        if "wrapping key" in lower:
            return "Wrapping key type"
        if "cryptographic actions" in lower or "field actions" in lower or "attribute actions" in lower or "字段级动作" in raw:
            return "Cryptographic actions / field actions"
        if "beacon" in lower or "searchable encryption" in lower:
            return "Searchable encryption / beacons"
        if "encryption context" in lower or "material description" in lower:
            return "Encryption context / material description"
        if "digital signatures" in lower or "algorithm suite" in lower or "默认/实际 algorithm" in raw:
            return "Algorithm suite / digital signatures"
        if "multi" in lower or "多租户" in raw:
            return "Multi-tenant key isolation"
        if "raw key" in lower:
            return "Raw keyring namespace/name"

    if service_id == "s3-encryption-client":
        if raw in {"Client-side object encryption for S3"}:
            return None
        if "cmm" == lower or "cryptographic materials manager" in lower:
            return "Cryptographic Materials Manager (CMM)"
        if "data key" in lower:
            return "Data key policy"
        if "encryption context" in lower:
            return "Encryption context"
        if "key commitment" in lower:
            return "Key commitment / commitment policy"
        if "keyring" in lower and "wrapping" not in lower:
            return "Keyring"
        if "wrapping key" in lower:
            return "Wrapping key"
        if "partial rsa" in lower or "key pair" in lower:
            return "Partial RSA key pair"
        if "metadata" in lower or "instruction files" in lower:
            return "Encryption metadata storage"
        if "对象内容加密算法" in raw or "对象加密算法" in raw:
            return "Object encryption algorithm"

    if service_id == "kms":
        if "external key store" in lower or "外部密钥库" in raw:
            return "External key store"
        if raw in {"KeySpec 密钥规格", "KeySpec：对称密钥", "对称加密 KeySpec"}:
            return "KeySpec: symmetric/default key"
        if raw in {"Origin", "密钥材料来源 Origin"}:
            return "Origin"
        if raw in {"导入密钥材料的过期时间", "Imported key material expiration 导入密钥材料过期策略"}:
            return "Imported key material expiration"
        if raw in {"按需轮换", "自动密钥轮换", "Automatic key rotation", "EnableKeyRotation"}:
            return "Key rotation"

    if service_id == "signer":
        if "platformid" in lower:
            return "platformId"
        if "signaturevalidityperiod" in lower:
            return "signatureValidityPeriod"
        if "signingmaterial.certificatearn" in lower:
            return "signingMaterial.certificateArn"

    return raw


def curate_config_items(items: list[dict]) -> list[dict]:
    curated: OrderedDict[tuple[str, str], dict] = OrderedDict()
    for item in items:
        sid = item["service_id"]
        canonical = canonical_config(sid, item["configuration_item"])
        if canonical is None:
            continue
        if (
            text(item["recommended_values"]) == text(item["configuration_item"])
            and text(item["risky_values"]) == text(item["configuration_item"])
        ):
            continue

        key = (sid, canonical)
        next_item = dict(item)
        next_item["configuration_item"] = canonical
        next_item["item_id"] = f"{sid}_{slug(canonical)}"

        if key not in curated:
            curated[key] = next_item
            continue

        existing = curated[key]
        for field in ["api_endpoint", "recommended_values", "risky_values", "risk_reason", "security_value_reason", "references"]:
            existing[field] = merge_field(existing.get(field, ""), next_item.get(field, ""))
        if existing.get("risk_level") != "高" and next_item.get("risk_level") == "高":
            existing["risk_level"] = "高"

    # Corrections from AWS official docs cross-checks.
    for item in curated.values():
        sid = item["service_id"]
        config = item["configuration_item"]
        if sid == "payment-cryptography" and config.startswith("KeyClass"):
            item["recommended_values"] = "SYMMETRIC_KEY 用于 PIN/CVV/MAC/数据加密等对称场景；ASYMMETRIC_KEY_PAIR、PRIVATE_KEY、PUBLIC_KEY 用于 TR-34、RSA/ECC、证书和密钥交换场景"
            item["risky_values"] = "KeyClass 与 KeyUsage/KeyAlgorithm 不匹配；忽略 PRIVATE_KEY/PUBLIC_KEY 等导入或证书场景所需类别"
            item["references"] = REFS["payment-cryptography"]
        if sid == "cloudfront" and config == "DistributionConfig.ViewerCertificate.MinimumProtocolVersion":
            item["api_endpoint"] = "UpdateDistribution / DistributionConfig.ViewerCertificate.MinimumProtocolVersion"
            item["recommended_values"] = "TLSv1.3_2025、TLSv1.2_2025、TLSv1.2_2021；按客户端兼容性选择尽可能新的 CloudFront security policy"
            item["risky_values"] = "SSLv3、TLSv1、TLSv1_2016、TLSv1.1_2016；TLSv1.2_2018/TLSv1.2_2019 仅作兼容回退"
            item["risk_reason"] = "该字段控制 viewer 到 CloudFront 的最低 TLS 协议版本和对应安全策略。旧值会允许过时协议或旧 cipher 组合，增加降级、兼容遗留弱客户端和合规风险。"
            item["security_value_reason"] = "TLSv1.3_2025、TLSv1.2_2025、TLSv1.2_2021 是当前更适合现代客户端的 CloudFront viewer security policy 选择；仍需结合客户端兼容性做灰度。"
            item["references"] = REFS["cloudfront"]
        if sid == "cloudfront" and config == "DistributionConfig.ViewerCertificate":
            item["api_endpoint"] = "DistributionConfig.ViewerCertificate.ACMCertificateArn / IAMCertificateId / CloudFrontDefaultCertificate"
            item["recommended_values"] = "自定义域名优先使用 ACM 证书；证书链、算法和域名匹配目标客户端；证书生命周期有续期和监控流程"
            item["risky_values"] = "自定义域名依赖默认证书、证书过期、导入证书无人管理、证书链或域名不匹配"
            item["risk_reason"] = "ViewerCertificate 决定 viewer HTTPS 使用的证书。证书来源、域名覆盖、链完整性和续期流程不当会导致连接失败、身份验证风险或私钥治理风险。"
            item["security_value_reason"] = "ACM 证书能与 CloudFront 集成并简化证书部署和续期治理；导入证书或 IAM 证书应只在明确需求下使用。"
            item["references"] = REFS["cloudfront"]
        if sid == "acm" and config == "RequestCertificate.KeyAlgorithm":
            item["recommended_values"] = "ACM 新申请证书支持 RSA_2048、EC_prime256v1、EC_secp384r1；默认 RSA_2048；ECDSA 需确认客户端兼容"
            item["risky_values"] = "把 RSA_1024、RSA_3072、RSA_4096、EC_secp521r1 当作 ACM 新申请证书可选项；它们属于导入证书相关算法范围"
            item["references"] = REFS["acm"]
        if sid == "acm" and config == "ValidationMethod":
            item["recommended_values"] = "DNS 验证；保留 ACM 创建的 CNAME 记录以支持自动续期"
            item["risky_values"] = "EMAIL 验证更依赖人工邮箱流程；误删 DNS 验证记录会影响续期"
        if sid == "acm" and config == "ExportCertificate.Passphrase":
            item["recommended_values"] = "使用强口令，并通过 fileb:// 文件传入导出私钥口令"
            item["risky_values"] = "弱口令、命令行明文传参、写入 shell 历史或口令管理薄弱"
        if sid == "acm" and config == "Options.Export":
            item["recommended_values"] = "默认 DISABLED；仅在确需跨平台部署证书私钥时启用导出"
            item["risky_values"] = "无业务必要时 ENABLED，会扩大私钥离开 ACM 边界后的暴露面"
        if sid == "acm" and config == "ImportCertificate / imported certificate lifecycle":
            item["recommended_values"] = "导入证书使用完整链、现代算法和受控私钥；建立过期监控与手动续期流程；可托管场景优先 ACM 签发证书"
            item["risky_values"] = "过期或弱算法导入证书、链不完整、私钥来源不明、误以为导入证书可由 ACM 自动续期"
        if sid == "acm" and config == "证书来源 / 证书类型":
            item["recommended_values"] = "公网证书优先 ACM 公有证书；内部 PKI 使用 ACM Private CA；外部 CA 证书仅在跨平台或既有信任链要求下导入"
            item["risky_values"] = "不区分 AMAZON_ISSUED、PRIVATE、IMPORTED 的生命周期差异；导入证书私钥和续期流程无人负责"
        if sid == "s3-encryption-client" and config == "Encryption context":
            item["recommended_values"] = "仅放非敏感、稳定、可审计的 name-value 对；使用对称 KMS wrapping key 时作为 KMS AAD；优先 US-ASCII 并在解密时保持一致"
            item["risky_values"] = "放入敏感信息；使用非 US-ASCII 导致可用性/兼容性问题；加解密 encryption context 不一致"
            item["references"] = REFS["s3-encryption-client"]
        if sid == "s3-encryption-client" and config == "Data key policy":
            item["recommended_values"] = "每个 S3 对象使用唯一 256-bit data key，由客户端生成和保护"
            item["risky_values"] = "外部实现复用固定 data key，或绕过客户端默认的每对象唯一 data key 设计"
            item["references"] = REFS["s3-encryption-client"]
        if sid == "s3-encryption-client" and config == "Wrapping key":
            item["recommended_values"] = "优先使用 AWS KMS wrapping key；确需 Raw AES-GCM/Raw RSA 时必须有成熟密钥管理和轮换流程"
            item["risky_values"] = "Raw AES-GCM/Raw RSA 自管不当；wrapping key 丢失或访问控制过宽；Go 3.x 场景误用不支持的 raw wrapping key"
            item["references"] = REFS["s3-encryption-client"]
        if sid == "s3-encryption-client" and config == "Encryption metadata storage":
            item["recommended_values"] = "默认使用对象 metadata 保存加密材料描述；若用 instruction files，需同步管理额外对象的权限、一致性和生命周期"
            item["risky_values"] = "instruction files 权限或生命周期与对象不一致；加密 metadata 丢失导致对象不可解密"
            item["references"] = REFS["s3-encryption-client"]
        if sid == "s3-encryption-client" and config == "Keyring":
            item["recommended_values"] = "使用内置 KMS keyring 或经过审计的自定义 keyring；keyring 应明确 wrapping keys"
            item["risky_values"] = "自定义 keyring 缺少审计、策略约束或 wrapping key 标识不清"
            item["references"] = REFS["s3-encryption-client"]
        if sid == "database-encryption-sdk" and config == "Searchable encryption / beacons":
            item["recommended_values"] = "仅在完成威胁建模后启用；标准 beacon 用于等值搜索，compound beacon 用于更复杂查询；为每个 beacon 配置二级索引"
            item["risky_values"] = "对既有数据直接启用并期望回填；用 DO_NOTHING 字段构建 beacon；忽略 beacon 泄露搜索模式和频率"
            item["references"] = REFS["database-encryption-sdk"]
        if sid == "database-encryption-sdk" and config == "Discovery filter":
            item["recommended_values"] = "discovery mode 下使用 account ID + partition 等严格 filter 限制可解密 key 范围"
            item["risky_values"] = "discovery mode 不加 filter，导致 SDK 可尝试任意能解开数据键的 KMS key"
            item["references"] = REFS["database-encryption-sdk"]
        if sid == "database-encryption-sdk" and config == "Keyring configuration":
            item["recommended_values"] = "优先 AWS KMS keyring；需要降低 KMS 调用时使用 AWS KMS Hierarchical keyring，并控制 branch key cache 生命周期"
            item["risky_values"] = "裸 discovery mode、raw keyring 管理不当、branch key cache 生命周期过宽或多租户材料隔离不足"
            item["references"] = REFS["database-encryption-sdk"]
        if sid == "database-encryption-sdk" and config == "Cryptographic actions / field actions":
            item["recommended_values"] = "敏感字段 ENCRYPT_AND_SIGN；需要明文索引的字段按需 SIGN_ONLY 或 SIGN_AND_INCLUDE_IN_ENCRYPTION_CONTEXT"
            item["risky_values"] = "敏感字段 DO_NOTHING；关键字段不签名；应保密字段只做 SIGN_ONLY"
            item["references"] = REFS["database-encryption-sdk"]
        if sid == "database-encryption-sdk" and config == "Strict mode / discovery mode":
            item["recommended_values"] = "优先 strict mode 并明确 KMS key ARN；必须 discovery mode 时加 discovery filter"
            item["risky_values"] = "无过滤 discovery mode；严格模式下 key 标识不具体导致解密范围或可用性问题"
            item["references"] = REFS["database-encryption-sdk"]
        if sid == "encryption-sdk" and config == "Commitment policy":
            item["recommended_values"] = "生产基线优先 RequireEncryptRequireDecrypt；迁移期可短期使用 RequireEncryptAllowDecrypt 并完成兼容性测试"
            item["risky_values"] = "长期使用 ForbidEncryptAllowDecrypt 或允许非 key commitment 算法套件"
            item["references"] = REFS["encryption-sdk"]
        if sid == "encryption-sdk" and config == "Keyring configuration":
            item["recommended_values"] = "优先 AWS KMS keyring；multi-keyring 只加入必要 wrapping keys；解密严格模式使用 key ARN"
            item["risky_values"] = "裸 discovery keyring、raw keyring 缺少治理、multi-keyring 范围过宽或接受过多解密 key"
            item["references"] = REFS["encryption-sdk"]
        if sid == "encryption-sdk" and config == "Encryption context":
            item["recommended_values"] = "放入非敏感、稳定、能在解密端重现的业务标识，例如 tenant、table、object type、purpose、version"
            item["risky_values"] = "放入敏感信息；上下文缺失领域隔离；解密端不校验关键 encryption context"
            item["references"] = REFS["encryption-sdk"]
        if sid in {"alb", "nlb"} and config in {"Listener SSL security policy", "TLS listener SSL security policy", "Listener.SslPolicy"}:
            item["recommended_values"] = "优先 ELBSecurityPolicy-TLS13-1-2-Res-PQ-2025-09；合规场景考虑 FIPS/PQ；按客户端能力逐步收紧"
            item["risky_values"] = "ELBSecurityPolicy-2016-08 或支持 TLS 1.0/1.1、CBC、非 FS cipher 的旧策略；CLI/CloudFormation/CDK 未显式指定导致默认旧策略"
            item["references"] = REFS[sid]

    result = sorted(curated.values(), key=lambda x: (x["service_id"], x["configuration_item"]))
    seen: dict[str, int] = {}
    for item in result:
        base = f"{item['service_id']}_{slug(item['configuration_item'])}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        item["item_id"] = base if count == 0 else f"{base}-{count + 1}"
    return result


def main():
    services = []
    for service_id, meta in SERVICE_META.items():
        service = {"service_id": service_id, **meta, "detail_path": f"#/services/{service_id}"}
        services.append(service)

    risk_types = [
        {
            "risk_id": risk_id,
            "risk_type": risk_type,
            "risk_description": description,
            "related_services": related,
            "typical_risky_values": risky,
            "recommended_direction": direction,
            "notes": "由当前目录 Excel 明细表整理，并按 AWS 官方文档核对关键可选值；总结性表格仅作辅助来源。",
        }
        for risk_id, risk_type, description, related, risky, direction in RISK_TYPES
    ]

    items: OrderedDict[str, dict] = OrderedDict()
    extract_generic_tables(items)
    extract_tls_policies(items)
    extract_cloudhsm_mechanisms(items)
    add_manual_items(items)

    config_items = curate_config_items(list(items.values()))
    for path, data in [
        (DATA_DIR / "services.json", services),
        (DATA_DIR / "risk_types.json", risk_types),
        (DATA_DIR / "config_items.json", config_items),
    ]:
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Wrote {path}: {len(data)} rows")


if __name__ == "__main__":
    main()
